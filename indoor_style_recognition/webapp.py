"""
FastAPI 网页应用：用于多人上传图片、识别、反馈和触发重训。
"""

import hmac
import json
import os
import sys
import threading
import zipfile
import shutil
import csv
from collections import Counter
from typing import Iterable
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from fastapi import FastAPI, File, Form, Request, Response, UploadFile
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from PIL import Image
from starlette.middleware.sessions import SessionMiddleware
from starlette.responses import RedirectResponse

sys.path.insert(0, str(Path(__file__).parent))

from config import (
    ADMIN_PASSWORD,
    CHECKPOINT_PATH,
    EXCEL_PATH,
    EXCLUDED_STYLE_CATEGORIES,
    PROCESSED_DIR,
    SESSION_SECRET,
    STYLE_DEFINITION_SHEET_NAME,
    TRAIN_IMAGE_ROOT,
)
from data_processor import DataProcessor
from feedback_loop import (
    append_feedback_log,
    append_pending_review_log,
    load_recent_feedback,
    load_pending_reviews,
    load_recent_wrong_feedback,
    save_feedback_image,
    save_pending_review_image,
    update_pending_review_status,
)
from inference import StylePredictor
from style_knowledge import STYLE_TEXT_COLUMNS, normalize_text
from train import Trainer


BASE_DIR = Path(__file__).parent
TEMPLATES_DIR = BASE_DIR / 'web_templates'
STATIC_DIR = BASE_DIR / 'web_static'
UPLOAD_DIR = Path(PROCESSED_DIR) / 'web_uploads'
TRAINING_STATUS_PATH = Path(PROCESSED_DIR) / 'training_status.json'
DATASET_IMAGE_DIR = Path(TRAIN_IMAGE_ROOT)
ASSET_UPLOAD_DIR = Path(PROCESSED_DIR) / 'asset_uploads'
ASSET_RESTORE_STATUS_PATH = Path(PROCESSED_DIR) / 'asset_restore_status.json'
REVIEW_QUEUE_DIR = Path(PROCESSED_DIR) / 'review_queue'
REVIEW_IMAGE_DIR = REVIEW_QUEUE_DIR / 'images'

STATIC_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
DATASET_IMAGE_DIR.mkdir(parents=True, exist_ok=True)
ASSET_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
REVIEW_IMAGE_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title='室内风格协作标注平台')
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    session_cookie='style_admin_session',
    same_site='lax',
    https_only=os.getenv('STYLE_SESSION_HTTPS_ONLY', '0') == '1',
)
app.mount('/static', StaticFiles(directory=str(STATIC_DIR)), name='static')
app.mount('/uploads', StaticFiles(directory=str(UPLOAD_DIR)), name='uploads')
app.mount('/dataset-images', StaticFiles(directory=str(DATASET_IMAGE_DIR)), name='dataset-images')
app.mount('/review-images', StaticFiles(directory=str(REVIEW_IMAGE_DIR)), name='review-images')
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

_predictor_cache = None
_predictor_init_error = ''
_training_thread = None
_training_lock = threading.Lock()
_asset_restore_thread = None
_asset_restore_lock = threading.Lock()
_style_definition_cache = {
    'excel_path': '',
    'mtime': None,
    'profiles': {},
}


def _safe_error_message(prefix, exc):
    return f'{prefix}: {exc}'


def invalidate_style_definition_cache():
    _style_definition_cache['excel_path'] = ''
    _style_definition_cache['mtime'] = None
    _style_definition_cache['profiles'] = {}


def normalize_style_name(style_name: str) -> str:
    return normalize_text(style_name or '')


def get_editable_style_profile_fields():
    fields = []
    for index, (_, alias) in enumerate(STYLE_TEXT_COLUMNS):
        if alias == '搜索关键词':
            continue
        fields.append({
            'alias': alias,
            'form_name': f'style_profile_{index}',
        })
    return fields


def ensure_style_directory(style_name: str) -> Path:
    normalized_style = normalize_style_name(style_name)
    if not normalized_style:
        raise ValueError('风格名称不能为空。')
    target_dir = Path(TRAIN_IMAGE_ROOT) / normalized_style
    target_dir.mkdir(parents=True, exist_ok=True)
    return target_dir


def ensure_style_definition_entry(style_name: str) -> bool:
    """确保 Excel 定义页里存在该风格，缺失时自动补一行空白定义。"""
    normalized_style = normalize_style_name(style_name)
    if not normalized_style:
        raise ValueError('风格名称不能为空。')

    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        raise FileNotFoundError(f'Excel 文件不存在：{excel_path}')

    workbook = load_workbook(excel_path)
    if STYLE_DEFINITION_SHEET_NAME not in workbook.sheetnames:
        raise KeyError(f'Excel 中缺少工作表：{STYLE_DEFINITION_SHEET_NAME}')

    sheet = workbook[STYLE_DEFINITION_SHEET_NAME]
    header_cells = [cell.value for cell in sheet[1]]
    header_map = {
        str(value).strip(): idx + 1
        for idx, value in enumerate(header_cells)
        if str(value).strip()
    }

    style_column = header_map.get('美学风格词')
    if not style_column:
        raise KeyError("风格定义页缺少 '美学风格词' 列，无法自动追加风格。")

    for row_idx in range(2, sheet.max_row + 1):
        existing_value = normalize_style_name(sheet.cell(row=row_idx, column=style_column).value)
        if existing_value == normalized_style:
            return False

    new_row_index = sheet.max_row + 1
    sheet.cell(row=new_row_index, column=style_column, value=normalized_style)

    # 为常用定义字段预留空白单元格，方便后续人工补全。
    for column_name, _ in STYLE_TEXT_COLUMNS:
        column_index = header_map.get(column_name)
        if column_index:
            sheet.cell(row=new_row_index, column=column_index, value='')

    workbook.save(excel_path)
    invalidate_style_definition_cache()
    return True


def save_style_definition_profile(style_name: str, profile_updates: dict[str, str]) -> tuple[bool, bool]:
    """把网页里的风格说明写回 Excel 定义页。"""
    normalized_style = normalize_style_name(style_name)
    if not normalized_style:
        raise ValueError('风格名称不能为空。')

    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        raise FileNotFoundError(f'Excel 文件不存在：{excel_path}')

    workbook = load_workbook(excel_path)
    if STYLE_DEFINITION_SHEET_NAME not in workbook.sheetnames:
        raise KeyError(f'Excel 中缺少工作表：{STYLE_DEFINITION_SHEET_NAME}')

    sheet = workbook[STYLE_DEFINITION_SHEET_NAME]
    header_cells = [cell.value for cell in sheet[1]]
    header_map = {
        str(value).strip(): idx + 1
        for idx, value in enumerate(header_cells)
        if str(value).strip()
    }

    style_column = header_map.get('美学风格词')
    if not style_column:
        raise KeyError("风格定义页缺少 '美学风格词' 列，无法保存风格说明。")

    row_index = None
    for current_row in range(2, sheet.max_row + 1):
        existing_value = normalize_style_name(sheet.cell(row=current_row, column=style_column).value)
        if existing_value == normalized_style:
            row_index = current_row
            break

    created_row = False
    if row_index is None:
        row_index = sheet.max_row + 1
        sheet.cell(row=row_index, column=style_column, value=normalized_style)
        created_row = True

    updated = False
    for column_name, alias in STYLE_TEXT_COLUMNS:
        if alias not in profile_updates:
            continue

        column_index = header_map.get(column_name)
        if not column_index:
            continue

        new_value = normalize_text(profile_updates.get(alias, ''))
        existing_value = normalize_text(sheet.cell(row=row_index, column=column_index).value)
        if existing_value == new_value:
            continue

        sheet.cell(row=row_index, column=column_index, value=new_value)
        updated = True

    workbook.save(excel_path)
    invalidate_style_definition_cache()
    return created_row, updated


def is_admin_authenticated(request: Request):
    return bool(request.session.get('is_admin_authenticated'))


def verify_admin_password(password: str):
    return hmac.compare_digest(password or '', ADMIN_PASSWORD)


def require_admin(request: Request):
    if is_admin_authenticated(request):
        return None
    return RedirectResponse(url='/admin/login', status_code=303)


def get_predictor():
    global _predictor_cache, _predictor_init_error
    if _predictor_cache is None:
        try:
            _predictor_cache = StylePredictor(validate_freshness=False)
            _predictor_init_error = ''
        except Exception as exc:
            _predictor_init_error = str(exc)
            raise
    return _predictor_cache


def refresh_predictor():
    global _predictor_cache, _predictor_init_error
    _predictor_cache = StylePredictor(validate_freshness=False)
    _predictor_init_error = ''
    return _predictor_cache


def get_predictor_if_available():
    try:
        return get_predictor()
    except Exception:
        return None


def predictor_artifacts_ready():
    required_paths = [
        Path(PROCESSED_DIR) / 'style_mapping.json',
        Path(PROCESSED_DIR) / 'style_profiles.json',
        Path(PROCESSED_DIR) / 'style_text_features.npz',
        Path(CHECKPOINT_PATH),
    ]
    return all(path.exists() for path in required_paths)


def dataset_image_url(path_value):
    path = Path(path_value)
    try:
        rel = path.resolve().relative_to(Path(TRAIN_IMAGE_ROOT).resolve())
        return f"/dataset-images/{rel.as_posix()}"
    except Exception:
        return ''


def upload_image_url(path_value):
    path = Path(path_value)
    try:
        rel = path.resolve().relative_to(UPLOAD_DIR.resolve())
        return f"/uploads/{rel.as_posix()}"
    except Exception:
        return ''


def review_image_url(path_value):
    path = Path(path_value)
    try:
        rel = path.resolve().relative_to(REVIEW_IMAGE_DIR.resolve())
        return f"/review-images/{rel.as_posix()}"
    except Exception:
        return ''


def list_available_styles():
    """合并当前模型风格与本地图库风格，供网页下拉选择。"""
    predictor = _predictor_cache
    style_set = set(predictor.style_to_label.keys()) if predictor is not None else set()

    image_root = Path(TRAIN_IMAGE_ROOT)
    if image_root.exists():
        for path in image_root.iterdir():
            if path.is_dir() and path.name.strip() and path.name not in EXCLUDED_STYLE_CATEGORIES:
                style_set.add(path.name.strip())

    return sorted(style_set)


def _human_size(num_bytes):
    value = float(num_bytes or 0)
    for unit in ('B', 'KB', 'MB', 'GB'):
        if value < 1024 or unit == 'GB':
            return f"{value:.0f} {unit}" if unit == 'B' else f"{value:.1f} {unit}"
        value /= 1024
    return f"{int(num_bytes)} B"


def load_style_definition_profiles():
    excel_path = Path(EXCEL_PATH)
    if not excel_path.exists():
        return {}

    mtime = excel_path.stat().st_mtime
    if (
        _style_definition_cache['excel_path'] == str(excel_path) and
        _style_definition_cache['mtime'] == mtime
    ):
        return _style_definition_cache['profiles']

    definition_df = pd.read_excel(excel_path, sheet_name=STYLE_DEFINITION_SHEET_NAME)
    if '美学风格词' not in definition_df.columns:
        return {}

    indexed = definition_df.copy()
    indexed['美学风格词'] = indexed['美学风格词'].apply(normalize_text)
    indexed = indexed[indexed['美学风格词'] != ''].drop_duplicates(subset=['美学风格词'], keep='first')

    profiles = {}
    for _, row in indexed.iterrows():
        style_name = row.get('美学风格词', '')
        if not style_name:
            continue
        profile = {'style': style_name}
        for column, alias in STYLE_TEXT_COLUMNS:
            profile[alias] = normalize_text(row[column]) if column in row.index else ''
        profiles[style_name] = profile

    _style_definition_cache['excel_path'] = str(excel_path)
    _style_definition_cache['mtime'] = mtime
    _style_definition_cache['profiles'] = profiles
    return profiles


def get_style_profile(style_name):
    if not style_name:
        return {}

    profile = load_style_definition_profiles().get(style_name)
    if profile:
        return profile

    return {
        'style': style_name,
        '定位': '',
        '定位依据': '',
        '定义': '',
        '色彩': '',
        '材料': '',
        '形态': '',
        '风格区别': '',
        '典型特征': '',
        '情绪氛围': '',
        '搜索关键词': '',
    }


def list_style_gallery_items(style_name):
    style_dir = Path(TRAIN_IMAGE_ROOT) / style_name
    image_extensions = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}
    root = Path(TRAIN_IMAGE_ROOT).resolve()
    items = []

    if not style_dir.exists():
        return items

    for image_path in sorted(style_dir.iterdir()):
        if not image_path.is_file() or image_path.suffix.lower() not in image_extensions:
            continue
        rel = image_path.resolve().relative_to(root)
        stat = image_path.stat()
        items.append({
            'name': image_path.name,
            'relative_path': rel.as_posix(),
            'url': f"/dataset-images/{rel.as_posix()}",
            'size_text': _human_size(stat.st_size),
            'updated_at': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
        })

    return items


def build_style_sidebar(search_query='', current_style=''):
    query = (search_query or '').strip().lower()
    styles = []
    for style_name in list_available_styles():
        image_count = len(list_style_gallery_items(style_name))
        if query and query not in style_name.lower():
            continue
        styles.append({
            'name': style_name,
            'image_count': image_count,
            'active': style_name == current_style,
        })
    return styles


def resolve_dataset_relative_path(relative_path):
    root = Path(TRAIN_IMAGE_ROOT).resolve()
    candidate = (root / relative_path).resolve()
    candidate.relative_to(root)
    return candidate


def unique_destination_path(directory: Path, filename: str):
    directory.mkdir(parents=True, exist_ok=True)
    stem = Path(filename).stem or 'image'
    suffix = Path(filename).suffix or '.jpg'
    candidate = directory / f"{stem}{suffix}"
    index = 1
    while candidate.exists():
        candidate = directory / f"{stem}_{index}{suffix}"
        index += 1
    return candidate


def build_style_gallery_context(request: Request, style_name='', search_query='', message='', error=''):
    sidebar = build_style_sidebar(search_query=search_query, current_style=style_name)
    selected_style = style_name or (sidebar[0]['name'] if sidebar else '')
    if selected_style and not any(item['name'] == selected_style for item in sidebar):
        sidebar = build_style_sidebar(search_query='', current_style=selected_style)

    selected_profile = get_style_profile(selected_style) if selected_style else {}
    selected_images = list_style_gallery_items(selected_style) if selected_style else []
    profile_field_descriptors = get_editable_style_profile_fields()
    for field in profile_field_descriptors:
        field['value'] = selected_profile.get(field['alias'], '')

    return build_context(
        skip_predictor_probe=True,
        admin_logged_in=is_admin_authenticated(request),
        message=message,
        error=error,
        style_search_query=search_query,
        style_sidebar=sidebar,
        selected_style=selected_style,
        selected_style_profile=selected_profile,
        selected_style_images=selected_images,
        selected_style_image_count=len(selected_images),
        style_profile_fields=[alias for _, alias in STYLE_TEXT_COLUMNS if alias != '搜索关键词'],
        style_profile_form_fields=profile_field_descriptors,
    )


def write_training_status(status, detail=''):
    payload = {
        'status': status,
        'detail': detail,
        'updated_at': __import__('datetime').datetime.now().isoformat(timespec='seconds'),
    }
    TRAINING_STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def write_asset_restore_status(status, detail=''):
    payload = {
        'status': status,
        'detail': detail,
        'updated_at': datetime.now().isoformat(timespec='seconds'),
    }
    ASSET_RESTORE_STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def load_asset_restore_status():
    if ASSET_RESTORE_STATUS_PATH.exists():
        return json.loads(ASSET_RESTORE_STATUS_PATH.read_text(encoding='utf-8'))
    return {
        'status': 'idle',
        'detail': '尚未开始上传云端资产。',
        'updated_at': '',
    }


def load_training_status():
    if TRAINING_STATUS_PATH.exists():
        return json.loads(TRAINING_STATUS_PATH.read_text(encoding='utf-8'))
    return {
        'status': 'idle',
        'detail': '尚未记录训练任务。',
        'updated_at': '',
    }


def load_recent_feedback_safe(limit=20):
    try:
        return load_recent_feedback(limit=limit)
    except Exception:
        return []


def load_recent_wrong_feedback_safe(limit=20):
    try:
        return load_recent_wrong_feedback(limit=limit)
    except Exception:
        return []


def load_pending_reviews_safe(limit=50):
    try:
        return load_pending_reviews(limit=limit, status='pending')
    except Exception:
        return []


UPLOAD_RANKING_WINDOWS = {
    'all': {'label': '全部时间', 'days': None},
    '7d': {'label': '近 7 天', 'days': 7},
    '30d': {'label': '近 30 天', 'days': 30},
}

UPLOAD_RANKING_STATUSES = {
    'all': '全部上传',
    'approved': '已通过',
    'pending': '待审核',
    'rejected': '已驳回',
}

TREND_ANALYSIS_WINDOWS = {
    '14d': {'label': '近 14 天', 'days': 14, 'bucket': 'day'},
    '30d': {'label': '近 30 天', 'days': 30, 'bucket': 'day'},
    '90d': {'label': '近 90 天', 'days': 90, 'bucket': 'week'},
    'all': {'label': '全部时间', 'days': None, 'bucket': 'month'},
}


def parse_iso_datetime(value):
    text = (value or '').strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def start_of_day(value: datetime) -> datetime:
    return datetime(value.year, value.month, value.day)


def start_of_week(value: datetime) -> datetime:
    day_start = start_of_day(value)
    return day_start - timedelta(days=day_start.weekday())


def start_of_month(value: datetime) -> datetime:
    return datetime(value.year, value.month, 1)


def next_month(value: datetime) -> datetime:
    if value.month == 12:
        return datetime(value.year + 1, 1, 1)
    return datetime(value.year, value.month + 1, 1)


def floor_bucket_start(value: datetime, bucket: str) -> datetime:
    if bucket == 'day':
        return start_of_day(value)
    if bucket == 'week':
        return start_of_week(value)
    return start_of_month(value)


def next_bucket_start(value: datetime, bucket: str) -> datetime:
    if bucket == 'day':
        return value + timedelta(days=1)
    if bucket == 'week':
        return value + timedelta(days=7)
    return next_month(value)


def format_bucket_label(bucket_start: datetime, bucket: str) -> tuple[str, str]:
    if bucket == 'day':
        return bucket_start.strftime('%m-%d'), bucket_start.strftime('%Y-%m-%d')
    if bucket == 'week':
        bucket_end = bucket_start + timedelta(days=6)
        return (
            f"{bucket_start.strftime('%m/%d')}-{bucket_end.strftime('%m/%d')}",
            f"{bucket_start.strftime('%Y-%m-%d')} 至 {bucket_end.strftime('%Y-%m-%d')}",
        )
    return bucket_start.strftime('%Y-%m'), bucket_start.strftime('%Y 年 %m 月')


def build_bucket_starts(range_start: datetime, range_end: datetime, bucket: str) -> list[datetime]:
    starts = []
    current = floor_bucket_start(range_start, bucket)
    end = floor_bucket_start(range_end, bucket)
    while current <= end:
        starts.append(current)
        current = next_bucket_start(current, bucket)
    return starts


def list_style_preview_images(style_name, limit=3):
    return [
        {'url': item['url'], 'name': item['name']}
        for item in list_style_gallery_items(style_name)[:limit]
    ]


def iter_trend_analysis_events():
    feedback_log_path = Path(PROCESSED_DIR) / 'feedback_log.csv'
    if feedback_log_path.exists():
        with open(feedback_log_path, 'r', encoding='utf-8', newline='') as handle:
            for row in csv.DictReader(handle):
                style_name = (row.get('chosen_style') or '').strip()
                submitted_at = parse_iso_datetime(row.get('timestamp'))
                if not style_name or submitted_at is None:
                    continue
                yield {
                    'style': style_name,
                    'status': 'approved',
                    'submitted_at': submitted_at,
                    'uploaded_at': submitted_at,
                    'activity_at': submitted_at,
                    'is_correct': row.get('is_correct') == '1',
                }

    pending_log_path = Path(PROCESSED_DIR) / 'review_queue' / 'pending_reviews.csv'
    if pending_log_path.exists():
        with open(pending_log_path, 'r', encoding='utf-8', newline='') as handle:
            for row in csv.DictReader(handle):
                style_name = (row.get('chosen_style') or '').strip()
                status = (row.get('status') or '').strip().lower()
                submitted_at = parse_iso_datetime(row.get('timestamp'))
                reviewed_at = parse_iso_datetime(row.get('reviewed_at'))
                if not style_name or status not in {'pending', 'rejected'}:
                    continue
                activity_at = reviewed_at if status == 'rejected' and reviewed_at is not None else submitted_at
                if activity_at is None:
                    continue
                yield {
                    'style': style_name,
                    'status': status,
                    'submitted_at': submitted_at or activity_at,
                    'uploaded_at': submitted_at or activity_at,
                    'activity_at': activity_at,
                    'is_correct': row.get('is_correct') == '1',
                }


def iter_upload_ranking_events():
    for event in iter_trend_analysis_events():
        yield {
            'style': event['style'],
            'status': event['status'],
            'uploaded_at': event['uploaded_at'],
            'activity_at': event['activity_at'],
        }


def build_upload_ranking_snapshot(top_n=12, window_key='all', status_filter='all'):
    selected_window = window_key if window_key in UPLOAD_RANKING_WINDOWS else 'all'
    selected_status = status_filter if status_filter in UPLOAD_RANKING_STATUSES else 'all'
    window_days = UPLOAD_RANKING_WINDOWS[selected_window]['days']
    now = datetime.now()
    range_start = now - timedelta(days=window_days) if window_days else None
    compare_days = window_days or 30
    compare_start = now - timedelta(days=compare_days)
    previous_start = compare_start - timedelta(days=compare_days)

    all_events = list(iter_upload_ranking_events())
    filtered_events = [
        event for event in all_events
        if (selected_status == 'all' or event['status'] == selected_status)
        and (range_start is None or event['activity_at'] >= range_start)
    ]

    counts = Counter(event['style'] for event in filtered_events)
    approved_counts = Counter(event['style'] for event in filtered_events if event['status'] == 'approved')
    pending_counts = Counter(event['style'] for event in filtered_events if event['status'] == 'pending')
    rejected_counts = Counter(event['style'] for event in filtered_events if event['status'] == 'rejected')
    library_counts = {style_name: len(list_style_gallery_items(style_name)) for style_name in counts}

    total_uploads = sum(counts.values())
    ranking_rows = []
    for index, (style_name, count) in enumerate(counts.most_common(), start=1):
        ratio = (count / total_uploads) if total_uploads else 0.0
        style_events = [event for event in filtered_events if event['style'] == style_name]
        latest_uploaded_at = max((event['uploaded_at'] for event in style_events), default=None)
        ranking_rows.append({
            'rank': index,
            'style': style_name,
            'count': count,
            'ratio': ratio,
            'ratio_percent': round(ratio * 100, 2),
            'approved_count': approved_counts.get(style_name, 0),
            'pending_count': pending_counts.get(style_name, 0),
            'rejected_count': rejected_counts.get(style_name, 0),
            'library_image_count': library_counts.get(style_name, 0),
            'latest_uploaded_at': latest_uploaded_at.strftime('%Y-%m-%d %H:%M') if latest_uploaded_at else '暂无',
            'preview_images': list_style_preview_images(style_name, limit=3),
        })

    top_rows = ranking_rows[:top_n]
    other_count = sum(row['count'] for row in ranking_rows[top_n:])
    other_ratio = (other_count / total_uploads) if total_uploads else 0.0
    chart_rows = [
        {
            'style': row['style'],
            'count': row['count'],
            'height_ratio': (row['count'] / top_rows[0]['count']) if top_rows and top_rows[0]['count'] else 0,
            'ratio_percent': row['ratio_percent'],
        }
        for row in top_rows
    ]
    if other_count:
        chart_rows.append({
            'style': '其他',
            'count': other_count,
            'height_ratio': (other_count / top_rows[0]['count']) if top_rows and top_rows[0]['count'] else 0,
            'ratio_percent': round(other_ratio * 100, 2),
        })

    current_compare_events = [
        event for event in all_events
        if (selected_status == 'all' or event['status'] == selected_status)
        and event['activity_at'] >= compare_start
    ]
    previous_compare_events = [
        event for event in all_events
        if (selected_status == 'all' or event['status'] == selected_status)
        and previous_start <= event['activity_at'] < compare_start
    ]
    current_compare_counts = Counter(event['style'] for event in current_compare_events)
    previous_compare_counts = Counter(event['style'] for event in previous_compare_events)
    growth_rows = []
    for style_name in set(current_compare_counts) | set(previous_compare_counts):
        current_count = current_compare_counts.get(style_name, 0)
        previous_count = previous_compare_counts.get(style_name, 0)
        delta = current_count - previous_count
        growth_rows.append({
            'style': style_name,
            'current_count': current_count,
            'previous_count': previous_count,
            'delta': delta,
            'delta_label': f"{delta:+d}",
            'preview_images': list_style_preview_images(style_name, limit=3),
        })
    growth_rows.sort(key=lambda row: (row['delta'], row['current_count'], row['style']), reverse=True)
    growth_rows = growth_rows[:6]

    quality_rows = []
    for row in ranking_rows[:6]:
        total_known = row['approved_count'] + row['pending_count'] + row['rejected_count']
        quality_rows.append({
            'style': row['style'],
            'total': total_known,
            'approved': row['approved_count'],
            'pending': row['pending_count'],
            'rejected': row['rejected_count'],
            'library_image_count': row['library_image_count'],
            'latest_uploaded_at': row['latest_uploaded_at'],
            'preview_images': row['preview_images'],
        })

    preview_rows = []
    for row in top_rows[:6]:
        if not row['preview_images']:
            continue
        preview_rows.append({
            'style': row['style'],
            'count': row['count'],
            'ratio_percent': row['ratio_percent'],
            'preview_images': row['preview_images'],
        })

    return {
        'has_data': total_uploads > 0,
        'selected_window': selected_window,
        'selected_status': selected_status,
        'window_options': [{'value': key, 'label': item['label']} for key, item in UPLOAD_RANKING_WINDOWS.items()],
        'status_options': [{'value': key, 'label': label} for key, label in UPLOAD_RANKING_STATUSES.items()],
        'window_label': UPLOAD_RANKING_WINDOWS[selected_window]['label'],
        'status_label': UPLOAD_RANKING_STATUSES[selected_status],
        'total_uploads': total_uploads,
        'approved_total': sum(approved_counts.values()),
        'pending_total': sum(pending_counts.values()),
        'rejected_total': sum(rejected_counts.values()),
        'style_total': len(counts),
        'top_rows': top_rows,
        'all_rows': ranking_rows,
        'chart_rows': chart_rows,
        'growth_rows': growth_rows,
        'quality_rows': quality_rows,
        'preview_rows': preview_rows,
        'compare_window_label': f"最近 {compare_days} 天 vs 更早 {compare_days} 天",
        'export_url': f"/upload-ranking/export?window={selected_window}&status={selected_status}",
    }


def build_trend_analysis_snapshot(window_key='30d'):
    selected_window = window_key if window_key in TREND_ANALYSIS_WINDOWS else '30d'
    window_config = TREND_ANALYSIS_WINDOWS[selected_window]
    bucket = window_config['bucket']
    window_days = window_config['days']
    now = datetime.now()
    range_start = start_of_day(now - timedelta(days=window_days - 1)) if window_days else None
    style_first_seen = {}
    all_events = list(iter_trend_analysis_events())

    for event in all_events:
        submitted_at = event.get('submitted_at')
        style_name = event.get('style', '')
        if submitted_at is None or not style_name:
            continue
        existing = style_first_seen.get(style_name)
        if existing is None or submitted_at < existing:
            style_first_seen[style_name] = submitted_at

    filtered_events = [
        event for event in all_events
        if event.get('submitted_at') is not None
        and (range_start is None or event['submitted_at'] >= range_start)
    ]

    if not filtered_events:
        return {
            'has_data': False,
            'selected_window': selected_window,
            'window_label': window_config['label'],
            'window_options': [{'value': key, 'label': item['label']} for key, item in TREND_ANALYSIS_WINDOWS.items()],
            'bucket_label': {'day': '按天', 'week': '按周', 'month': '按月'}[bucket],
            'compare_days': min(window_days or 60, 30),
            'active_days': 0,
            'summary_cards': [],
            'timeline_rows': [],
            'surging_rows': [],
            'cooling_rows': [],
            'correction_rows': [],
            'emerging_rows': [],
            'insight_cards': [],
            'dominant_style': None,
            'peak_bucket': None,
            'preview_stats': [],
            'headline': '',
            'summary': '',
            'watch_items': [],
            'compact_timeline_rows': [],
            'signal_ribbon': [],
            'hero_highlights': [],
        }

    effective_range_start = range_start or min(event['submitted_at'] for event in filtered_events)
    bucket_starts = build_bucket_starts(effective_range_start, now, bucket)
    timeline_map = {
        bucket_start: {
            'total': 0,
            'approved': 0,
            'pending': 0,
            'rejected': 0,
            'corrected': 0,
        }
        for bucket_start in bucket_starts
    }

    style_counts = Counter()
    style_corrected_counts = Counter()
    style_status_counts = {}
    active_days = set()

    for event in filtered_events:
        bucket_start = floor_bucket_start(event['submitted_at'], bucket)
        bucket_row = timeline_map.get(bucket_start)
        if bucket_row is None:
            continue

        style_name = event['style']
        status = event['status']
        is_correct = event.get('is_correct', False)
        active_days.add(event['submitted_at'].date())
        style_counts[style_name] += 1
        if not is_correct:
            style_corrected_counts[style_name] += 1
        status_counter = style_status_counts.setdefault(style_name, Counter())
        status_counter[status] += 1

        bucket_row['total'] += 1
        bucket_row[status] += 1
        if not is_correct:
            bucket_row['corrected'] += 1

    total_events = sum(style_counts.values())
    corrected_total = sum(style_corrected_counts.values())
    approved_total = sum(counts.get('approved', 0) for counts in style_status_counts.values())
    pending_total = sum(counts.get('pending', 0) for counts in style_status_counts.values())
    rejected_total = sum(counts.get('rejected', 0) for counts in style_status_counts.values())
    active_style_total = len(style_counts)
    correction_rate_percent = round((corrected_total / total_events) * 100, 1) if total_events else 0.0

    timeline_max = max((row['total'] for row in timeline_map.values()), default=0)
    timeline_rows = []
    for bucket_start in bucket_starts:
        row = timeline_map[bucket_start]
        label, full_label = format_bucket_label(bucket_start, bucket)
        total = row['total']
        timeline_rows.append({
            'label': label,
            'full_label': full_label,
            'total': total,
            'approved_count': row['approved'],
            'pending_count': row['pending'],
            'rejected_count': row['rejected'],
            'corrected_count': row['corrected'],
            'height_percent': max(10.0, round((total / timeline_max) * 100, 2)) if total and timeline_max else 0.0,
            'approved_share': round((row['approved'] / total) * 100, 2) if total else 0.0,
            'pending_share': round((row['pending'] / total) * 100, 2) if total else 0.0,
            'rejected_share': round((row['rejected'] / total) * 100, 2) if total else 0.0,
            'is_peak': bool(total and total == timeline_max),
        })

    dominant_style = None
    if total_events and style_counts:
        style_name, style_count = style_counts.most_common(1)[0]
        dominant_style = {
            'style': style_name,
            'count': style_count,
            'ratio_percent': round((style_count / total_events) * 100, 1),
        }

    peak_bucket = next((row for row in timeline_rows if row['is_peak']), None)

    compare_days = min(window_days or 60, 30)
    current_start = start_of_day(now - timedelta(days=compare_days - 1))
    previous_start = start_of_day(current_start - timedelta(days=compare_days))
    current_counts = Counter(
        event['style'] for event in all_events
        if event.get('submitted_at') is not None and event['submitted_at'] >= current_start
    )
    previous_counts = Counter(
        event['style'] for event in all_events
        if event.get('submitted_at') is not None and previous_start <= event['submitted_at'] < current_start
    )
    momentum_rows = []
    for style_name in set(current_counts) | set(previous_counts):
        current_count = current_counts.get(style_name, 0)
        previous_count = previous_counts.get(style_name, 0)
        delta = current_count - previous_count
        momentum_rows.append({
            'style': style_name,
            'current_count': current_count,
            'previous_count': previous_count,
            'delta': delta,
            'delta_label': f"{delta:+d}",
            'preview_images': list_style_preview_images(style_name, limit=3),
        })
    momentum_rows.sort(key=lambda row: (row['delta'], row['current_count'], row['style']), reverse=True)
    surging_rows = [row for row in momentum_rows if row['delta'] > 0][:5]
    cooling_rows = sorted(
        [row for row in momentum_rows if row['delta'] < 0],
        key=lambda row: (row['delta'], row['current_count'], row['style'])
    )[:5]

    correction_rows = []
    for style_name, total in style_counts.items():
        if total < 2:
            continue
        corrected = style_corrected_counts.get(style_name, 0)
        status_counts = style_status_counts.get(style_name, Counter())
        correction_rows.append({
            'style': style_name,
            'total': total,
            'corrected': corrected,
            'correction_rate_percent': round((corrected / total) * 100, 1),
            'approved': status_counts.get('approved', 0),
            'pending': status_counts.get('pending', 0),
            'rejected': status_counts.get('rejected', 0),
        })
    correction_rows.sort(
        key=lambda row: (row['correction_rate_percent'], row['corrected'], row['total'], row['style']),
        reverse=True,
    )
    correction_rows = correction_rows[:6]

    if selected_window == 'all':
        new_style_cutoff = start_of_day(now - timedelta(days=30 - 1))
        emerging_label = '近 30 天首次出现'
    else:
        new_style_cutoff = effective_range_start
        emerging_label = f"{window_config['label']} 内首次出现"

    emerging_rows = []
    for style_name, first_seen in style_first_seen.items():
        if first_seen < new_style_cutoff:
            continue
        emerging_rows.append({
            'style': style_name,
            'first_seen_label': first_seen.strftime('%Y-%m-%d'),
            'total': style_counts.get(style_name, 0),
            'preview_images': list_style_preview_images(style_name, limit=3),
        })
    emerging_rows.sort(key=lambda row: (row['first_seen_label'], row['total'], row['style']), reverse=True)
    emerging_rows = emerging_rows[:6]

    new_style_total = len(emerging_rows)
    summary_cards = [
        {
            'eyebrow': 'Signals',
            'value': str(total_events),
            'label': '时间范围内的总反馈量',
        },
        {
            'eyebrow': 'Correction Rate',
            'value': f'{correction_rate_percent:.1f}%',
            'label': '需要人工纠错的占比',
        },
        {
            'eyebrow': 'Active Styles',
            'value': str(active_style_total),
            'label': '被用户触达的风格数',
        },
        {
            'eyebrow': 'New Styles',
            'value': str(new_style_total),
            'label': emerging_label,
        },
    ]

    insight_cards = []
    if dominant_style:
        insight_cards.append({
            'eyebrow': 'Dominant Style',
            'title': dominant_style['style'],
            'detail': f"当前时间范围内占比 {dominant_style['ratio_percent']}%，共 {dominant_style['count']} 次。",
        })
    if peak_bucket:
        insight_cards.append({
            'eyebrow': 'Peak Window',
            'title': peak_bucket['full_label'],
            'detail': f"这一段时间最活跃，共收到 {peak_bucket['total']} 次反馈。",
        })
    insight_cards.append({
        'eyebrow': 'Review Flow',
        'title': f'通过 {approved_total} / 待审 {pending_total} / 驳回 {rejected_total}',
        'detail': '可以同时观察用户偏好变化和审核压力变化。',
    })

    preview_stats = []
    if dominant_style:
        preview_stats.append({
            'label': '当前最热风格',
            'value': dominant_style['style'],
            'meta': f"{dominant_style['ratio_percent']}% · {dominant_style['count']} 次",
        })
    preview_stats.append({
        'label': '人工纠错占比',
        'value': f'{correction_rate_percent:.1f}%',
        'meta': window_config['label'],
    })
    if peak_bucket:
        preview_stats.append({
            'label': '最活跃时段',
            'value': peak_bucket['label'],
            'meta': f"{peak_bucket['total']} 次反馈",
        })

    headline = '最近的用户偏好正在分散流动。'
    if dominant_style is not None:
        headline = f"{dominant_style['style']} 是当前最强势的主流风格。"
    if surging_rows and dominant_style and surging_rows[0]['style'] != dominant_style['style']:
        headline = f"{surging_rows[0]['style']} 正在快速升温，而 {dominant_style['style']} 仍保持主流优势。"
    elif surging_rows:
        headline = f"{surging_rows[0]['style']} 不只是热门，还在继续升温。"

    summary_parts = [
        f"在 {window_config['label']} 内共记录 {total_events} 次反馈，覆盖 {active_style_total} 个风格。",
        f"人工纠错占比 {correction_rate_percent:.1f}%，审核流转为通过 {approved_total}、待审 {pending_total}、驳回 {rejected_total}。",
    ]
    if peak_bucket is not None:
        summary_parts.append(f"最活跃的时间段是 {peak_bucket['full_label']}。")
    summary = ' '.join(summary_parts)

    watch_items = []
    if surging_rows:
        top_surging = surging_rows[0]
        watch_items.append(f"升温关注：{top_surging['style']} 较上一周期 {top_surging['delta_label']}。")
    if correction_rows:
        top_correction = correction_rows[0]
        watch_items.append(f"纠错关注：{top_correction['style']} 的人工纠错占比达到 {top_correction['correction_rate_percent']}%。")
    if emerging_rows:
        top_emerging = emerging_rows[0]
        watch_items.append(f"新增关注：{top_emerging['style']} 于 {top_emerging['first_seen_label']} 首次出现。")
    if not watch_items and dominant_style is not None:
        watch_items.append(f"主流关注：{dominant_style['style']} 当前占比 {dominant_style['ratio_percent']}%。")

    compact_timeline_source = timeline_rows[-12:]
    compact_timeline_max = max((row['total'] for row in compact_timeline_source), default=0)
    compact_timeline_rows = []
    for row in compact_timeline_source:
        compact_timeline_rows.append({
            'label': row['label'],
            'full_label': row['full_label'],
            'total': row['total'],
            'corrected_count': row['corrected_count'],
            'height_percent': max(14.0, round((row['total'] / compact_timeline_max) * 100, 2)) if row['total'] and compact_timeline_max else 0.0,
            'is_peak': bool(row['total'] and row['total'] == compact_timeline_max),
        })

    signal_ribbon = []
    if dominant_style is not None:
        signal_ribbon.append({
            'tone': 'neutral',
            'label': '主流风格',
            'value': dominant_style['style'],
            'meta': f"{dominant_style['ratio_percent']}%",
        })
    if surging_rows:
        signal_ribbon.append({
            'tone': 'up',
            'label': '升温最快',
            'value': surging_rows[0]['style'],
            'meta': surging_rows[0]['delta_label'],
        })
    if correction_rows:
        signal_ribbon.append({
            'tone': 'warn',
            'label': '纠错关注',
            'value': correction_rows[0]['style'],
            'meta': f"{correction_rows[0]['correction_rate_percent']}%",
        })
    elif peak_bucket is not None:
        signal_ribbon.append({
            'tone': 'neutral',
            'label': '活跃时段',
            'value': peak_bucket['label'],
            'meta': f"{peak_bucket['total']} 次",
        })

    hero_highlights = [
        {
            'label': '反馈信号',
            'value': str(total_events),
            'meta': window_config['label'],
        },
        {
            'label': '主流风格',
            'value': dominant_style['style'] if dominant_style else '暂无',
            'meta': (
                f"{dominant_style['ratio_percent']}% 占比"
                if dominant_style else '等待更多样本'
            ),
        },
        {
            'label': '纠错压力',
            'value': f'{correction_rate_percent:.1f}%',
            'meta': '人工纠错占比',
        },
    ]

    return {
        'has_data': True,
        'selected_window': selected_window,
        'window_label': window_config['label'],
        'window_options': [{'value': key, 'label': item['label']} for key, item in TREND_ANALYSIS_WINDOWS.items()],
        'bucket_label': {'day': '按天', 'week': '按周', 'month': '按月'}[bucket],
        'compare_days': compare_days,
        'summary_cards': summary_cards,
        'timeline_rows': timeline_rows,
        'surging_rows': surging_rows,
        'cooling_rows': cooling_rows,
        'correction_rows': correction_rows,
        'emerging_rows': emerging_rows,
        'emerging_label': emerging_label,
        'insight_cards': insight_cards,
        'dominant_style': dominant_style,
        'peak_bucket': peak_bucket,
        'preview_stats': preview_stats,
        'active_days': len(active_days),
        'headline': headline,
        'summary': summary,
        'watch_items': watch_items,
        'compact_timeline_rows': compact_timeline_rows,
        'signal_ribbon': signal_ribbon,
        'hero_highlights': hero_highlights,
    }


def run_incremental_training_workflow():
    """后台执行数据处理与快速微调。"""
    try:
        write_training_status('running', '正在重建数据集...')
        processor = DataProcessor()
        processor.process_data()

        write_training_status('running', '正在基于最近反馈做快速微调...')
        trainer = Trainer(num_workers_override=0)
        trainer.run_incremental()

        write_training_status('running', '正在重新加载最新模型...')
        refresh_predictor()
        write_training_status('completed', '后台快速微调已完成，最新模型已加载。')
    except Exception as e:
        write_training_status('failed', f'后台训练失败: {e}')


def start_background_incremental_training():
    """如果当前没有训练任务，则启动一个后台训练线程。"""
    global _training_thread
    with _training_lock:
        if _training_thread is not None and _training_thread.is_alive():
            return False

        _training_thread = threading.Thread(
            target=run_incremental_training_workflow,
            name='style-incremental-training',
            daemon=True,
        )
        _training_thread.start()
        return True


def build_training_snapshot():
    processed_path = Path(PROCESSED_DIR) / 'processed_data.csv'
    mapping_path = Path(PROCESSED_DIR) / 'style_mapping.json'
    model_path = Path(CHECKPOINT_PATH)

    snapshot = {
        'processed_exists': processed_path.exists(),
        'mapping_exists': mapping_path.exists(),
        'model_exists': model_path.exists(),
        'processed_rows': 0,
        'processed_styles': 0,
        'model_updated_at': model_path.stat().st_mtime if model_path.exists() else None,
        'training_status': load_training_status(),
        'feedback_total': len(load_recent_feedback_safe(limit=100000)),
        'feedback_wrong_total': len(load_recent_wrong_feedback_safe(limit=100000)),
        'snapshot_error': '',
    }

    if processed_path.exists():
        try:
            import pandas as pd
            df = pd.read_csv(processed_path, encoding='utf-8')
            snapshot['processed_rows'] = int(len(df))
            snapshot['processed_styles'] = int(df['style'].nunique()) if 'style' in df.columns else 0
        except Exception as exc:
            snapshot['snapshot_error'] = _safe_error_message('processed 数据读取失败', exc)

    if snapshot['model_updated_at'] is not None:
        from datetime import datetime
        snapshot['model_updated_at_text'] = datetime.fromtimestamp(snapshot['model_updated_at']).strftime('%Y-%m-%d %H:%M:%S')
    else:
        snapshot['model_updated_at_text'] = '未生成'

    return snapshot


def build_context(skip_predictor_probe=False, **extra):
    asset_restore_status = load_asset_restore_status()
    should_probe_predictor = (
        not skip_predictor_probe and
        asset_restore_status.get('status') != 'running'
    )
    predictor = get_predictor_if_available() if should_probe_predictor else _predictor_cache
    recent_wrong_feedback = []
    for row in load_recent_wrong_feedback_safe(limit=12):
        row = dict(row)
        row['saved_url'] = dataset_image_url(row.get('saved_path', ''))
        recent_wrong_feedback.append(row)

    pending_reviews = []
    for row in load_pending_reviews_safe(limit=20):
        row = dict(row)
        row['pending_url'] = review_image_url(row.get('pending_path', ''))
        pending_reviews.append(row)

    base = {
        'styles': list_available_styles(),
        'recent_feedback': load_recent_feedback_safe(limit=15),
        'recent_wrong_feedback': recent_wrong_feedback,
        'pending_reviews': pending_reviews,
        'training_snapshot': build_training_snapshot(),
        'predictor_is_stale': predictor.is_stale if predictor is not None else False,
        'predictor_stale_reason': predictor.stale_reason if predictor is not None else '',
        'predictor_available': predictor is not None,
        'predictor_error': _predictor_init_error,
        'asset_restore_status': asset_restore_status,
        'message': '',
        'error': '',
        'result': None,
        'asset_uploads': {
            'processed_bundle': (ASSET_UPLOAD_DIR / 'render_processed_bundle.zip').exists(),
            'model_bundle': (ASSET_UPLOAD_DIR / 'render_model_bundle.zip').exists(),
            'excel_bundle': (ASSET_UPLOAD_DIR / 'render_excel_bundle.zip').exists(),
        },
        'admin_logged_in': False,
    }
    base.update(extra)
    return base


def _extract_zip_bytes(binary: bytes, destination: Path):
    if destination.exists():
        shutil.rmtree(destination)
    destination.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(BytesIO(binary)) as archive:
        archive.extractall(destination)


def _copy_tree_contents(source: Path, destination: Path):
    destination.mkdir(parents=True, exist_ok=True)
    for item in source.iterdir():
        target = destination / item.name
        if item.is_dir():
            if target.exists():
                shutil.rmtree(target)
            shutil.copytree(item, target)
        else:
            shutil.copy2(item, target)


def _extract_zip_file(zip_path: Path, destination: Path):
    if destination.exists():
        shutil.rmtree(destination)
    destination.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as archive:
        archive.extractall(destination)


def _find_single_excel_file(root: Path) -> Path:
    excel_files = sorted(
        path for path in root.rglob('*')
        if path.is_file() and path.suffix.lower() in {'.xlsx', '.xls'}
    )
    if not excel_files:
        raise FileNotFoundError('excel zip 中缺少 xlsx/xls 文件')
    return excel_files[0]


def _save_uploaded_file(upload: UploadFile, destination: Path):
    destination.parent.mkdir(parents=True, exist_ok=True)
    upload.file.seek(0)
    with destination.open('wb') as handle:
        shutil.copyfileobj(upload.file, handle)


def _apply_uploaded_assets():
    processed_zip = ASSET_UPLOAD_DIR / 'render_processed_bundle.zip'
    model_zip = ASSET_UPLOAD_DIR / 'render_model_bundle.zip'
    excel_zip = ASSET_UPLOAD_DIR / 'render_excel_bundle.zip'

    missing = [
        label for label, path in (
            ('processed zip', processed_zip),
            ('model zip', model_zip),
            ('excel zip', excel_zip),
        ) if not path.exists()
    ]
    if missing:
        raise FileNotFoundError(f"还缺这些文件：{', '.join(missing)}")

    tmp_root = Path(PROCESSED_DIR) / 'asset_imports'
    tmp_root.mkdir(parents=True, exist_ok=True)
    processed_unpack = tmp_root / 'processed_unpack'
    model_unpack = tmp_root / 'model_unpack'
    excel_unpack = tmp_root / 'excel_unpack'

    write_asset_restore_status('running', '正在解压 processed / model / excel 三个资产包...')
    _extract_zip_file(processed_zip, processed_unpack)
    _extract_zip_file(model_zip, model_unpack)
    _extract_zip_file(excel_zip, excel_unpack)

    processed_src = processed_unpack / 'indoor_style_recognition' / 'data' / 'processed'
    model_src = model_unpack / 'indoor_style_recognition' / 'models' / 'best_model.pth'
    excel_src = _find_single_excel_file(excel_unpack)

    if not processed_src.exists():
        raise FileNotFoundError('processed zip 中缺少 indoor_style_recognition/data/processed')
    if not model_src.exists():
        raise FileNotFoundError('model zip 中缺少 indoor_style_recognition/models/best_model.pth')

    write_asset_restore_status('running', '正在把处理文件、模型和 Excel 复制到 Render 磁盘...')
    _copy_tree_contents(processed_src, Path(PROCESSED_DIR))
    Path(CHECKPOINT_PATH).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(model_src, Path(CHECKPOINT_PATH))
    shutil.copy2(excel_src, Path(EXCEL_PATH))

    write_asset_restore_status('running', '文件复制完成，正在重新加载模型...')
    refresh_predictor()
    write_asset_restore_status('completed', '云端资产已恢复完成，模型已重新加载。')


def run_background_asset_restore():
    try:
        _apply_uploaded_assets()
    except Exception as exc:
        write_asset_restore_status('failed', f'云端资产恢复失败：{exc}')


def start_background_asset_restore():
    global _asset_restore_thread
    with _asset_restore_lock:
        if _asset_restore_thread is not None and _asset_restore_thread.is_alive():
            return False

        _asset_restore_thread = threading.Thread(
            target=run_background_asset_restore,
            name='asset-restore-worker',
            daemon=True,
        )
        _asset_restore_thread.start()
        return True


def render_asset_restore_page(request: Request, message='', error='', status_code=200):
    return templates.TemplateResponse(
        request=request,
        name='asset_restore.html',
        context=build_context(
            skip_predictor_probe=True,
            admin_logged_in=is_admin_authenticated(request),
            message=message,
            error=error,
        ),
        status_code=status_code,
    )


@app.get('/', response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        request=request,
        name='index.html',
        context=build_context(
            admin_logged_in=is_admin_authenticated(request),
            trend_analysis_preview=build_trend_analysis_snapshot(window_key='30d'),
        ),
    )


@app.head('/')
async def index_head():
    return Response(status_code=200)


@app.get('/styles', response_class=HTMLResponse)
async def style_gallery(
    request: Request,
    style: str = '',
    q: str = '',
):
    return templates.TemplateResponse(
        request=request,
        name='style_gallery.html',
        context=build_style_gallery_context(
            request=request,
            style_name=style.strip(),
            search_query=q.strip(),
        ),
    )


@app.get('/upload-ranking', response_class=HTMLResponse)
async def upload_ranking(request: Request, window: str = 'all', status: str = 'all'):
    return templates.TemplateResponse(
        request=request,
        name='upload_ranking.html',
        context=build_context(
            skip_predictor_probe=True,
            admin_logged_in=is_admin_authenticated(request),
            upload_ranking=build_upload_ranking_snapshot(window_key=window, status_filter=status),
        ),
    )


@app.get('/trend-analysis', response_class=HTMLResponse)
async def trend_analysis(request: Request, window: str = '30d'):
    return templates.TemplateResponse(
        request=request,
        name='trend_analysis.html',
        context=build_context(
            skip_predictor_probe=True,
            admin_logged_in=is_admin_authenticated(request),
            trend_analysis=build_trend_analysis_snapshot(window_key=window),
        ),
    )


@app.get('/upload-assets', response_class=HTMLResponse)
async def asset_restore_page(request: Request):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect
    return render_asset_restore_page(request=request)


@app.get('/upload-ranking/export')
async def export_upload_ranking(window: str = 'all', status: str = 'all'):
    snapshot = build_upload_ranking_snapshot(top_n=2000, window_key=window, status_filter=status)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'rank',
        'style',
        'count',
        'ratio_percent',
        'approved_count',
        'pending_count',
        'rejected_count',
        'library_image_count',
        'latest_uploaded_at',
        'window',
        'status',
    ])
    for row in snapshot['all_rows']:
        writer.writerow([
            row['rank'],
            row['style'],
            row['count'],
            row['ratio_percent'],
            row['approved_count'],
            row['pending_count'],
            row['rejected_count'],
            row['library_image_count'],
            row['latest_uploaded_at'],
            snapshot['window_label'],
            snapshot['status_label'],
        ])

    filename = f"upload-ranking-{snapshot['selected_window']}-{snapshot['selected_status']}.csv"
    return Response(
        content='\ufeff' + output.getvalue(),
        media_type='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@app.post('/styles/upload', response_class=HTMLResponse)
async def upload_style_images(
    request: Request,
    style_name: str = Form(...),
    search_query: str = Form(''),
    image_files: list[UploadFile] = File(...),
):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect

    saved_count = 0
    image_extensions = {'.jpg', '.jpeg', '.png', '.webp', '.bmp'}
    normalized_style = normalize_style_name(style_name)

    try:
        if not normalized_style:
            raise ValueError('请先选择要上传到哪个风格。')

        target_dir = ensure_style_directory(normalized_style)
        created_excel_entry = ensure_style_definition_entry(normalized_style)

        for upload in image_files:
            suffix = Path(upload.filename or '').suffix.lower()
            if suffix not in image_extensions:
                continue
            destination = unique_destination_path(target_dir, upload.filename or f'upload{suffix}')
            _save_uploaded_file(upload, destination)
            saved_count += 1

        if saved_count == 0:
            raise ValueError('没有上传成功的图片，请确认文件格式为 jpg/png/webp/bmp。')

        message = f'已上传 {saved_count} 张图片到风格「{normalized_style}」。'
        if created_excel_entry:
            message += ' 已自动把该风格追加到 Excel 定义页。'

        context = build_style_gallery_context(
            request=request,
            style_name=normalized_style,
            search_query=search_query.strip(),
            message=message,
        )
    except Exception as exc:
        context = build_style_gallery_context(
            request=request,
            style_name=normalized_style,
            search_query=search_query.strip(),
            error=f'上传图片失败：{exc}',
        )

    return templates.TemplateResponse(request=request, name='style_gallery.html', context=context)


@app.post('/styles/create', response_class=HTMLResponse)
async def create_style(
    request: Request,
    new_style_name: str = Form(...),
    search_query: str = Form(''),
):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect

    normalized_style = normalize_style_name(new_style_name)

    try:
        if not normalized_style:
            raise ValueError('请输入新风格名称。')

        ensure_style_directory(normalized_style)
        created_excel_entry = ensure_style_definition_entry(normalized_style)
        message = f'已创建风格「{normalized_style}」。'
        if created_excel_entry:
            message += ' 已自动追加到 Excel 定义页。'
        else:
            message += ' Excel 定义页里原本就有这个风格。'

        context = build_style_gallery_context(
            request=request,
            style_name=normalized_style,
            search_query=search_query.strip(),
            message=message,
        )
    except Exception as exc:
        context = build_style_gallery_context(
            request=request,
            search_query=search_query.strip(),
            error=f'创建风格失败：{exc}',
        )

    return templates.TemplateResponse(request=request, name='style_gallery.html', context=context)


@app.post('/styles/profile/update', response_class=HTMLResponse)
async def update_style_profile(request: Request):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect

    form = await request.form()
    style_name = normalize_style_name(form.get('style_name', ''))
    search_query = normalize_text(form.get('search_query', ''))

    try:
        if not style_name:
            raise ValueError('请先选择要编辑的风格。')

        profile_updates = {}
        for field in get_editable_style_profile_fields():
            profile_updates[field['alias']] = normalize_text(form.get(field['form_name'], ''))

        created_excel_entry, updated = save_style_definition_profile(style_name, profile_updates)
        message = f'已保存风格「{style_name}」的说明到 Excel。'
        if created_excel_entry:
            message += ' 已自动创建该风格的定义行。'
        elif not updated:
            message = f'风格「{style_name}」的说明没有变化。'

        context = build_style_gallery_context(
            request=request,
            style_name=style_name,
            search_query=search_query,
            message=message,
        )
    except Exception as exc:
        context = build_style_gallery_context(
            request=request,
            style_name=style_name,
            search_query=search_query,
            error=f'保存风格说明失败：{exc}',
        )

    return templates.TemplateResponse(request=request, name='style_gallery.html', context=context)


@app.post('/styles/delete', response_class=HTMLResponse)
async def delete_style_image(
    request: Request,
    style_name: str = Form(...),
    search_query: str = Form(''),
    relative_path: str = Form(...),
):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect

    try:
        target_path = resolve_dataset_relative_path(relative_path)
        target_path.unlink(missing_ok=True)
        context = build_style_gallery_context(
            request=request,
            style_name=style_name.strip(),
            search_query=search_query.strip(),
            message=f'已删除图片：{Path(relative_path).name}',
        )
    except Exception as exc:
        context = build_style_gallery_context(
            request=request,
            style_name=style_name.strip(),
            search_query=search_query.strip(),
            error=f'删除图片失败：{exc}',
        )

    return templates.TemplateResponse(request=request, name='style_gallery.html', context=context)


@app.post('/styles/move', response_class=HTMLResponse)
async def move_style_image(
    request: Request,
    style_name: str = Form(...),
    target_style: str = Form(...),
    search_query: str = Form(''),
    relative_path: str = Form(...),
):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect

    try:
        normalized_target_style = normalize_style_name(target_style)
        if not normalized_target_style:
            raise ValueError('请选择目标风格。')

        source_path = resolve_dataset_relative_path(relative_path)
        destination_dir = ensure_style_directory(normalized_target_style)
        created_excel_entry = ensure_style_definition_entry(normalized_target_style)
        destination = unique_destination_path(destination_dir, source_path.name)
        shutil.move(str(source_path), str(destination))
        message = f'已把图片移动到风格「{normalized_target_style}」。'
        if created_excel_entry:
            message += ' 已自动把该风格追加到 Excel 定义页。'
        context = build_style_gallery_context(
            request=request,
            style_name=normalized_target_style,
            search_query=search_query.strip(),
            message=message,
        )
    except Exception as exc:
        context = build_style_gallery_context(
            request=request,
            style_name=style_name.strip(),
            search_query=search_query.strip(),
            error=f'更改风格失败：{exc}',
        )

    return templates.TemplateResponse(request=request, name='style_gallery.html', context=context)


@app.get('/admin/login', response_class=HTMLResponse)
async def admin_login(request: Request):
    if is_admin_authenticated(request):
        return RedirectResponse(url='/admin/reviews', status_code=303)
    return templates.TemplateResponse(
        request=request,
        name='admin_login.html',
        context=build_context(
            skip_predictor_probe=True,
            admin_logged_in=is_admin_authenticated(request),
        ),
    )


@app.post('/admin/login', response_class=HTMLResponse)
async def admin_login_submit(
    request: Request,
    password: str = Form(...),
):
    if verify_admin_password(password):
        request.session['is_admin_authenticated'] = True
        return RedirectResponse(url='/admin/reviews', status_code=303)

    return templates.TemplateResponse(
        request=request,
        name='admin_login.html',
        context=build_context(
            skip_predictor_probe=True,
            admin_logged_in=is_admin_authenticated(request),
            error='管理员密码不正确，请重试。',
        ),
        status_code=401,
    )


@app.post('/admin/logout')
async def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse(url='/admin/login', status_code=303)


@app.get('/admin/reviews', response_class=HTMLResponse)
async def admin_reviews(request: Request):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect
    return templates.TemplateResponse(
        request=request,
        name='admin_reviews.html',
        context=build_context(skip_predictor_probe=True, admin_logged_in=True),
    )


@app.head('/admin/reviews')
async def admin_reviews_head():
    return Response(status_code=200)


@app.post('/predict', response_class=HTMLResponse)
async def predict(
    request: Request,
    image_file: UploadFile = File(...),
):
    predictor = get_predictor_if_available()
    if predictor is None:
        return templates.TemplateResponse(
            request=request,
            name='index.html',
            context=build_context(
                error='当前云端模型尚未准备好，网站已经上线，但还缺少 processed 数据或模型文件。',
                admin_logged_in=is_admin_authenticated(request),
            ),
        )
    suffix = Path(image_file.filename or 'upload.jpg').suffix or '.jpg'
    upload_path = UPLOAD_DIR / f"upload_{Path(image_file.filename or 'image').stem}{suffix}"
    binary = await image_file.read()
    upload_path.write_bytes(binary)

    image = Image.open(upload_path).convert('RGB')
    result = predictor.predict(image, top_k=3, threshold=0.1)
    result['upload_path'] = str(upload_path.resolve())
    result['upload_name'] = image_file.filename or upload_path.name
    result['upload_url'] = upload_image_url(result['upload_path'])

    return templates.TemplateResponse(
        request=request,
        name='index.html',
        context=build_context(
            result=result,
            message='识别完成，请确认结果或提交正确风格。',
            admin_logged_in=is_admin_authenticated(request),
        ),
    )


@app.post('/feedback', response_class=HTMLResponse)
async def feedback(
    request: Request,
    upload_path: str = Form(...),
    predicted_style: str = Form(...),
    pred_1_style: str = Form(''),
    pred_1_confidence: str = Form(''),
    pred_2_style: str = Form(''),
    pred_2_confidence: str = Form(''),
    pred_3_style: str = Form(''),
    pred_3_confidence: str = Form(''),
    feedback_action: str = Form(...),
    chosen_style: str = Form(''),
    new_style_name: str = Form(''),
    retrain_now: str = Form('n'),
):
    predictor = get_predictor_if_available()
    if predictor is None:
        return templates.TemplateResponse(
            request=request,
            name='index.html',
            context=build_context(
                error='当前模型尚未准备好，暂时无法提交识别反馈。',
                admin_logged_in=is_admin_authenticated(request),
            ),
        )
    image_path = Path(upload_path)
    if not image_path.exists():
        return templates.TemplateResponse(
            request=request,
            name='index.html',
            context=build_context(
                error='上传的临时图片不存在，请重新上传。',
                admin_logged_in=is_admin_authenticated(request),
            ),
        )

    predictions = [
        {'style': pred_1_style, 'confidence': float(pred_1_confidence or 0.0)},
        {'style': pred_2_style, 'confidence': float(pred_2_confidence or 0.0)},
        {'style': pred_3_style, 'confidence': float(pred_3_confidence or 0.0)},
    ]

    is_correct = feedback_action == 'correct'
    final_style = predicted_style if is_correct else (new_style_name.strip() or chosen_style.strip())
    if not final_style:
        return templates.TemplateResponse(
            request=request,
            name='index.html',
            context=build_context(
                error='请选择正确风格，或者输入一个新的风格名称后再提交反馈。',
                admin_logged_in=is_admin_authenticated(request),
            ),
        )

    image = predictor.load_image(str(image_path))
    pending_path = save_pending_review_image(
        predictor=predictor,
        image_input=str(image_path),
        style_name=final_style,
        image=image,
    )
    review_id = append_pending_review_log(
        image_input=str(image_path),
        predictions=predictions,
        chosen_style=final_style,
        pending_path=pending_path,
        is_correct=is_correct,
    )

    message = f'反馈已进入待审核队列，审核编号 {review_id}。审核通过后才会加入训练库。'
    write_training_status('idle', '最近一次反馈已进入待审核队列，尚未触发重训。')

    if image_path.exists():
        image_path.unlink()

    return templates.TemplateResponse(
        request=request,
        name='index.html',
        context=build_context(
            message=message,
            admin_logged_in=is_admin_authenticated(request),
        ),
    )


@app.post('/review-feedback', response_class=HTMLResponse)
async def review_feedback(
    request: Request,
    review_id: str = Form(...),
    review_action: str = Form(...),
    review_note: str = Form(''),
):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect
    predictor = get_predictor_if_available()
    if predictor is None:
        return templates.TemplateResponse(
            request=request,
            name='admin_reviews.html',
            context=build_context(error='当前模型尚未准备好，暂时无法处理审核。', admin_logged_in=True),
        )

    pending_rows = load_pending_reviews_safe(limit=500)
    row = next((item for item in pending_rows if item.get('review_id') == review_id), None)
    if row is None:
        return templates.TemplateResponse(
            request=request,
            name='admin_reviews.html',
            context=build_context(error=f'未找到待审核记录：{review_id}', admin_logged_in=True),
        )

    pending_path = Path(row.get('pending_path', ''))
    if not pending_path.exists():
        return templates.TemplateResponse(
            request=request,
            name='admin_reviews.html',
            context=build_context(error=f'待审核图片不存在：{pending_path}', admin_logged_in=True),
        )

    if review_action == 'approve':
        image = predictor.load_image(str(pending_path))
        approved_style = normalize_style_name(row.get('chosen_style', '').strip())
        ensure_style_directory(approved_style)
        created_excel_entry = ensure_style_definition_entry(approved_style)
        saved_path = save_feedback_image(
            predictor=predictor,
            image_input=str(pending_path),
            style_name=approved_style,
            image=image,
        )
        predictions = []
        for idx in range(1, 4):
            predictions.append({
                'style': row.get(f'pred_{idx}_style', ''),
                'confidence': float(row.get(f'pred_{idx}_confidence', 0.0) or 0.0),
            })
        append_feedback_log(
            image_input=row.get('image_source', ''),
            predictions=predictions,
            chosen_style=row.get('chosen_style', ''),
            saved_path=saved_path,
            is_correct=row.get('is_correct') == '1',
        )
        update_pending_review_status(
            review_id=review_id,
            status='approved',
            review_note=review_note,
            approved_saved_path=saved_path,
        )
        pending_path.unlink(missing_ok=True)
        message = f'审核已通过，图片已加入训练库：{saved_path}'
        if created_excel_entry:
            message += ' 已自动把该风格追加到 Excel 定义页。'
    elif review_action == 'reject':
        update_pending_review_status(
            review_id=review_id,
            status='rejected',
            review_note=review_note,
            approved_saved_path='',
        )
        pending_path.unlink(missing_ok=True)
        message = f'审核已驳回：{review_id}'
    else:
        return templates.TemplateResponse(
            request=request,
            name='admin_reviews.html',
            context=build_context(error='未知审核动作，请重试。', admin_logged_in=True),
        )

    return templates.TemplateResponse(
        request=request,
        name='admin_reviews.html',
        context=build_context(message=message, admin_logged_in=True),
    )


@app.post('/refresh-model', response_class=HTMLResponse)
async def refresh_model(request: Request):
    try:
        refresh_predictor()
        message = '模型已重新加载。'
        error = ''
    except Exception as exc:
        message = ''
        error = f'模型重新加载失败：{exc}'
    return templates.TemplateResponse(
        request=request,
        name='index.html',
        context=build_context(
            message=message,
            error=error,
            admin_logged_in=is_admin_authenticated(request),
        ),
    )


@app.post('/upload-asset-bundle', response_class=HTMLResponse)
async def upload_asset_bundle(
    request: Request,
    bundle_kind: str = Form(...),
    bundle_file: UploadFile = File(...),
):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect
    try:
        bundle_map = {
            'processed': 'render_processed_bundle.zip',
            'model': 'render_model_bundle.zip',
            'excel': 'render_excel_bundle.zip',
        }
        if bundle_kind not in bundle_map:
            raise ValueError('未知资产类型，请重新选择。')
        if not (bundle_file.filename or '').lower().endswith('.zip'):
            raise ValueError('请上传 zip 文件。')

        destination = ASSET_UPLOAD_DIR / bundle_map[bundle_kind]
        _save_uploaded_file(bundle_file, destination)
        write_asset_restore_status('idle', f'{bundle_kind} 资产包已上传，等待应用。')
        label_map = {
            'processed': 'processed zip',
            'model': 'model zip',
            'excel': 'excel zip',
        }
        return render_asset_restore_page(
            request=request,
            message=f'{label_map[bundle_kind]} 已上传到云端暂存区。',
        )
    except Exception as exc:
        return render_asset_restore_page(
            request=request,
            error=f'资产上传失败：{exc}',
        )


@app.post('/upload-assets', response_class=HTMLResponse)
async def upload_assets(request: Request):
    redirect = require_admin(request)
    if redirect is not None:
        return redirect
    try:
        started = start_background_asset_restore()
        message = (
            '云端资产恢复任务已经启动，请等待 10-60 秒后刷新页面。'
            if started else
            '云端资产恢复任务已经在后台运行，请稍候刷新页面查看结果。'
        )
        return render_asset_restore_page(
            request=request,
            message=message,
        )
    except Exception as exc:
        return render_asset_restore_page(
            request=request,
            error=f'云端资产上传失败：{exc}',
        )


@app.get('/training-status', response_class=HTMLResponse)
async def training_status(request: Request):
    return templates.TemplateResponse(
        request=request,
        name='training_status.html',
        context=build_context(skip_predictor_probe=True),
    )


@app.head('/training-status')
async def training_status_head():
    return Response(status_code=200)


@app.get('/healthz')
async def healthz():
    return {
        'ok': True,
        'predictor_ready': _predictor_cache is not None,
        'predictor_files_ready': predictor_artifacts_ready(),
        'num_styles': _predictor_cache.num_classes if _predictor_cache is not None else 0,
        'predictor_error': _predictor_init_error,
    }
